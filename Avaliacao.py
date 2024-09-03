import tkinter as tk
from tkinter import Tk, Label, Button
from openpyxl import Workbook, load_workbook

arquivo_excel = "Itens_cadastrados.xlsx"  

entrada_item = None
entrada_valor_compra = None
entrada_porcentagem = None

def salvar_dados():  

    global entrada_item, entrada_valor_compra, entrada_porcentagem

    nome_item = entrada_item.get()
    valor_compra = int(entrada_valor_compra.get())
    porcentagem_lucro = int(entrada_porcentagem.get())
    valor_venda = (valor_compra*(porcentagem_lucro/100)) + valor_compra

    try:

        excel = load_workbook(arquivo_excel)
        pagina = excel.active

    except FileNotFoundError:  

        excel = Workbook()
        pagina = excel.active
        pagina.append(["Nome do Item","Valor do Item","Porcentagem de Venda","Valor de Venda"])

    pagina.append([nome_item, valor_compra, porcentagem_lucro, valor_venda])
    excel.save(arquivo_excel)

    print(f"Novos itens adicionados com sucesso ao arquivo '{arquivo_excel}'.")

def abrir_janela():  #Cria a janela principal da aplicação

    global entrada_item, entrada_valor_compra, entrada_porcentagem

    nome = str(input("Digite seu nome: "))

    janela = Tk()
    
    #Define o tamanho da janela e o título
    janela.geometry("550x250")
    janela.title("Avaliação - Laboratório de Programação (Sistema de cadastros)")

    #Aba de Cadastro de Produtos
    #Cria e posiciona um rótulo (Label) para o campo de entrada do item
    texto_item = Label(janela, text="Digite o item que deseja cadastrar:")
    #O método (grid) é utilizado para posicionar widgets em uma grade (ou tabela)
    #com linhas e colunas. Cada widget é colocado em uma célula específica dessa grade.
    texto_item.grid(column=0, row=0, padx=10, pady=10)

    entrada_item = tk.Entry(janela, width=50)
    entrada_item.grid(column=1, row=0, padx=10, pady=10)  # padx=10, pady=10 adiciona uma margem de 10 pixels ao redor do widget.

    texto_valor_compra = Label(janela, text="Valor de Compra:")
    texto_valor_compra.grid(column=0, row=2, padx=10, pady=10)

    #Cria e posiciona um campo de entrada (Entry) para o valor de compra
    entrada_valor_compra = tk.Entry(janela, width=50)
    entrada_valor_compra.grid(column=1, row=2, padx=10, pady=10)

    texto_porcentagem = Label(janela, text="Porcentagem de lucro(%):")
    texto_porcentagem.grid(column=0, row=3, padx=10, pady=10)

    entrada_porcentagem = tk.Entry(janela, width=50)
    entrada_porcentagem.grid(column=1, row=3, padx=10, pady=10)

    #Cria e posiciona um botão (Button) para salvar as informações
    botao_salvar = Button(janela, text="Salvar", command=salvar_dados)
    botao_salvar.grid(column=1, row=4, padx=10, pady=20)

    #Cria e posiciona uma mensagem de boas vindas com nome definido (Simulando login)
    boas_vindas = Label(janela, text=f" Bem Vindo(a) - {nome}")
    boas_vindas.grid(column=0, row=4, padx=10, pady=10)

    janela.config(bg="lightblue")  # Define a cor de fundo da janela

    #Define o tamanho máximo e mínimo da janela
    janela.maxsize(550, 250)
    janela.minsize(550, 250)

    janela.mainloop()

def buscar_item():

    try:

        excel = load_workbook(arquivo_excel)
        pagina = excel.active

    except FileNotFoundError:

        print("Arquivo Excel não encontrado. Certifique-se de cadastrar itens antes de buscar.")
        exit()

    item_busca = input("Digite o nome do item que deseja buscar: ").strip()
    item_encontrado = False

    for linha in pagina.iter_rows(min_row=2):
        
        if linha[0].value == item_busca:  # Acessa o valor da célula corretamente

            menuzinho = f"""
            Nome do Item: {linha[0].value}
            Valor do Item: R${linha[1].value}
            Porcentagem de lucro: {linha[2].value}%
            Valor de Venda: R${linha[3].value}
            """

            print(menuzinho)

            item_encontrado = True
            break

    if not item_encontrado:
        print("Item não encontrado.")


menu = f"""

--------------- Menu principal ---------------

1 - Cadastramento de Itens
2 - Busca de Produtos
3 - Sair

"""

while True:

    print(menu)
    escolha = int(input("Escolha uma opção: "))

    match escolha:

        case 1:
            abrir_janela()

        case 2:
            buscar_item()

        case 3:
            break

        case _:
            print("\nDigite um valor válido! \n")


print(f"\n\nObrigado por usar nosso sistema :)\n\n")

